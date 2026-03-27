"""
gnat.viz.tabular
====================

Tabular views of workspace objects — terminal tables, HTML, CSV, and Excel.

Targets
-------
* **Terminal** — ANSI-colored tables via ``rich`` (falls back to plain text).
* **Jupyter** — ``display(HTML(...))`` rendered inline.
* **Browser** — standalone self-contained HTML file with sortable/filterable columns.
* **CSV** — plain comma-separated, ready for Excel or any data tool.
* **Excel / Power BI** — ``.xlsx`` via ``openpyxl`` with auto-column widths,
  freeze-panes, and a styled header row.  Drop the file into Power BI Desktop
  as a data source — every sheet maps to a Power BI table.

Usage::

    from gnat.viz import TabularView

    view = TabularView(workspace)

    # Terminal — print directly
    view.show()
    view.show(stix_type="indicator", sort_by="confidence", top=50)

    # Save to various formats
    view.to_html("report.html")
    view.to_csv("indicators.csv")
    view.to_excel("workspace.xlsx")   # one sheet per STIX type

    # Jupyter
    view.display()

    # In the CLI: gnat viz table --workspace apt28 --output html --file report.html
"""

from __future__ import annotations

import csv
import io
import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, TYPE_CHECKING

if TYPE_CHECKING:
    from gnat.context.workspace import Workspace
    from gnat.orm.base import STIXBase

logger = logging.getLogger(__name__)

# ── STIX type → display columns ────────────────────────────────────────────

_COLUMNS: Dict[str, List[str]] = {
    "indicator":     ["id", "name", "indicator_types", "pattern", "confidence",
                      "x_tlp", "x_rf_risk_score", "created"],
    "malware":       ["id", "name", "malware_types", "is_family", "confidence",
                      "x_tlp", "created"],
    "vulnerability": ["id", "name", "description", "x_cvss_score",
                      "x_published", "confidence", "created"],
    "threat-actor":  ["id", "name", "threat_actor_types", "confidence",
                      "x_tlp", "created"],
    "attack-pattern": ["id", "name", "description", "confidence", "created"],
    "relationship":  ["id", "relationship_type", "source_ref", "target_ref",
                      "x_enrichment_source", "created"],
    "_default":      ["id", "type", "name", "confidence", "created"],
}

# ── STIX type → terminal color ─────────────────────────────────────────────

_TYPE_COLORS: Dict[str, str] = {
    "indicator":      "cyan",
    "malware":        "red",
    "vulnerability":  "yellow",
    "threat-actor":   "magenta",
    "attack-pattern": "blue",
    "relationship":   "green",
}

_TLP_COLORS: Dict[str, str] = {
    "white": "white",
    "green": "green",
    "amber": "yellow",
    "red":   "red",
}


def _coerce(val: Any) -> str:
    """Flatten any value to a display string."""
    if val is None:
        return ""
    if isinstance(val, list):
        return ", ".join(str(v) for v in val)
    if isinstance(val, bool):
        return "yes" if val else "no"
    if isinstance(val, float):
        return f"{val:.1f}"
    return str(val)


def _get_field(obj: "STIXBase", field: str) -> Any:
    """Safely retrieve a field from a STIXBase object."""
    if field == "type":
        return obj.stix_type
    if hasattr(obj, field):
        return getattr(obj, field, None)
    return obj._properties.get(field)


def _to_rows(objects: List["STIXBase"], columns: List[str]) -> List[Dict[str, str]]:
    return [{col: _coerce(_get_field(obj, col)) for col in columns}
            for obj in objects]


class TabularView:
    """
    Renders workspace objects as tables across multiple output formats.

    Parameters
    ----------
    workspace : Workspace
        The workspace to visualize.
    default_top : int
        Default maximum rows to show (terminal only).  Default 100.

    Examples
    --------
    >>> view = TabularView(workspace)
    >>> view.show()
    >>> view.show(stix_type="indicator", sort_by="confidence", top=25)
    >>> view.to_excel("analysis.xlsx")
    """

    def __init__(self, workspace: "Workspace", default_top: int = 100):
        self._ws       = workspace
        self._top      = default_top

    # ── Public API ──────────────────────────────────────────────────────────

    def show(
        self,
        stix_type: Optional[str] = None,
        sort_by: Optional[str] = "created",
        top: Optional[int] = None,
        fields: Optional[List[str]] = None,
    ) -> None:
        """
        Print a formatted table to the terminal.

        Uses ``rich`` if available for ANSI color, falls back to plain ASCII.

        Parameters
        ----------
        stix_type : str, optional
            Filter to only this STIX type.  If omitted all types are shown
            grouped by type.
        sort_by : str, optional
            Field name to sort by.  Default ``"created"``.
        top : int, optional
            Maximum rows.  Defaults to ``self.default_top``.
        fields : list of str, optional
            Override displayed columns.
        """
        limit = top or self._top
        groups = self._group_objects(stix_type)

        try:
            self._show_rich(groups, sort_by, limit, fields)
        except ImportError:
            self._show_plain(groups, sort_by, limit, fields)

    def display(
        self,
        stix_type: Optional[str] = None,
        sort_by: Optional[str] = "created",
        top: int = 200,
    ) -> None:
        """
        Render inline in a Jupyter notebook via ``IPython.display``.

        Falls back to :meth:`show` outside of Jupyter.
        """
        try:
            from IPython.display import display, HTML  # type: ignore
            html = self._build_html(self._group_objects(stix_type), sort_by, top)
            display(HTML(html))
        except ImportError:
            self.show(stix_type=stix_type, sort_by=sort_by, top=top)

    def to_html(
        self,
        path: Optional[str] = None,
        stix_type: Optional[str] = None,
        sort_by: Optional[str] = "created",
        top: int = 5000,
    ) -> str:
        """
        Generate a self-contained, sortable/filterable HTML report.

        Parameters
        ----------
        path : str, optional
            If provided, write the HTML to this file path.
        stix_type : str, optional
            Filter to one STIX type.  If omitted all types are included as
            tabbed sections.

        Returns
        -------
        str
            The complete HTML string.
        """
        html = self._build_html(self._group_objects(stix_type), sort_by, top)
        if path:
            Path(path).write_text(html, encoding="utf-8")
            logger.info("TabularView: HTML written to %s", path)
        return html

    def to_csv(
        self,
        path: str,
        stix_type: Optional[str] = None,
        sort_by: Optional[str] = "created",
        top: int = 100_000,
    ) -> None:
        """
        Write workspace objects to a CSV file.

        Parameters
        ----------
        path : str
            Output file path.
        stix_type : str, optional
            If omitted all objects are written; ``type`` is included as a column.
        """
        groups = self._group_objects(stix_type)
        rows: List[Dict[str, str]] = []
        all_cols: List[str] = []

        for stype, objs in groups.items():
            cols = _COLUMNS.get(stype, _COLUMNS["_default"])
            if "type" not in cols:
                cols = ["type"] + cols
            sorted_objs = self._sort(objs, sort_by)[:top]
            for row in _to_rows(sorted_objs, cols):
                rows.append(row)
            for c in cols:
                if c not in all_cols:
                    all_cols.append(c)

        if not rows:
            logger.warning("TabularView: no objects to export")
            return

        with open(path, "w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.DictWriter(fh, fieldnames=all_cols, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
        logger.info("TabularView: CSV written to %s (%d rows)", path, len(rows))

    def to_excel(
        self,
        path: str,
        sort_by: Optional[str] = "created",
        top: int = 100_000,
    ) -> None:
        """
        Write workspace objects to an Excel workbook — one sheet per STIX type.

        The workbook is Power BI-compatible: one sheet per type, typed columns,
        freeze panes, auto-width, header styling.

        Requires ``openpyxl``: ``pip install "gnat[viz]"``.

        Parameters
        ----------
        path : str
            Output ``.xlsx`` file path.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export: pip install 'gnat[viz]'"
            )

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        # ── Summary sheet ──────────────────────────────────────────────────
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["GNAT Workspace Export"])
        ws_summary.append(["Workspace", self._ws.name])
        ws_summary.append(["Exported", datetime.now(timezone.utc).isoformat()])
        ws_summary.append(["Total objects", len(self._ws)])
        ws_summary["A1"].font = Font(bold=True, size=14)

        # ── One sheet per STIX type ────────────────────────────────────────
        header_fill  = PatternFill("solid", fgColor="1a73e8")
        header_font  = Font(bold=True, color="FFFFFF")
        alt_fill     = PatternFill("solid", fgColor="EEF2FF")

        for stype, objs in self._group_objects().items():
            if not objs:
                continue
            cols = _COLUMNS.get(stype, _COLUMNS["_default"])
            sorted_objs = self._sort(objs, sort_by)[:top]
            rows = _to_rows(sorted_objs, cols)

            # Sheet name max 31 chars, no special chars
            sheet_name = stype.replace("-", " ").title()[:31]
            ws = wb.create_sheet(sheet_name)

            # Header row
            ws.append(cols)
            for cell in ws[1]:
                cell.font  = header_font
                cell.fill  = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Data rows with alternating fill
            for i, row in enumerate(rows, start=2):
                ws.append([row.get(c, "") for c in cols])
                if i % 2 == 0:
                    for cell in ws[i]:
                        cell.fill = alt_fill

            # Freeze header, auto-width columns
            ws.freeze_panes = "A2"
            for col_idx, col_name in enumerate(cols, start=1):
                col_letter = get_column_letter(col_idx)
                max_len    = max(
                    (len(str(r.get(col_name, ""))) for r in rows),
                    default=0
                )
                ws.column_dimensions[col_letter].width = min(
                    max(len(col_name), max_len) + 4, 60
                )

        wb.save(path)
        logger.info("TabularView: Excel written to %s", path)

    def to_dataframe(self, stix_type: Optional[str] = None):
        """
        Return a ``pandas.DataFrame`` of workspace objects.

        Requires ``pandas``: ``pip install pandas``.

        Returns
        -------
        pd.DataFrame
        """
        try:
            import pandas as pd
        except ImportError:
            raise ImportError("pandas is required: pip install pandas")

        groups = self._group_objects(stix_type)
        all_rows = []
        for stype, objs in groups.items():
            cols = _COLUMNS.get(stype, _COLUMNS["_default"])
            for obj in objs:
                row = {col: _get_field(obj, col) for col in cols}
                row["_stix_type"] = stype
                all_rows.append(row)
        return pd.DataFrame(all_rows)

    # ── Internals ────────────────────────────────────────────────────────────

    def _group_objects(
        self, stix_type: Optional[str] = None
    ) -> Dict[str, List["STIXBase"]]:
        """Group workspace objects by stix_type."""
        result: Dict[str, List["STIXBase"]] = {}
        for obj in self._ws.objects.values():
            if stix_type and obj.stix_type != stix_type:
                continue
            result.setdefault(obj.stix_type, []).append(obj)
        return dict(sorted(result.items()))

    @staticmethod
    def _sort(objs: List["STIXBase"], sort_by: Optional[str]) -> List["STIXBase"]:
        if not sort_by:
            return objs
        def _key(obj):
            v = _get_field(obj, sort_by)
            if v is None:
                return ""
            if isinstance(v, (int, float)):
                return -v  # descending for numbers
            return str(v)
        return sorted(objs, key=_key)

    def _show_rich(self, groups, sort_by, limit, fields) -> None:
        from rich.console import Console
        from rich.table import Table
        from rich import box as rich_box

        console = Console()
        console.print(f"\n[bold]Workspace:[/bold] [cyan]{self._ws.name}[/cyan]  "
                      f"[dim]{len(self._ws)} objects[/dim]\n")

        for stype, objs in groups.items():
            cols  = fields or _COLUMNS.get(stype, _COLUMNS["_default"])
            color = _TYPE_COLORS.get(stype, "white")
            table = Table(
                title=f"[{color}]{stype}[/{color}]  ({len(objs)} objects)",
                box=rich_box.ROUNDED,
                show_header=True,
                header_style="bold",
                row_styles=["", "dim"],
            )
            for col in cols:
                table.add_column(col, overflow="fold",
                                 max_width=60 if col in ("pattern", "description") else 40)
            for obj in self._sort(objs, sort_by)[:limit]:
                row = [_coerce(_get_field(obj, c)) for c in cols]
                table.add_row(*row)
            console.print(table)
            console.print()

    def _show_plain(self, groups, sort_by, limit, fields) -> None:
        for stype, objs in groups.items():
            cols = fields or _COLUMNS.get(stype, _COLUMNS["_default"])
            sorted_objs = self._sort(objs, sort_by)[:limit]
            rows = _to_rows(sorted_objs, cols)
            widths = {c: max(len(c), max((len(r[c]) for r in rows), default=0))
                      for c in cols}
            print(f"\n── {stype} ({len(objs)}) " + "─" * 40)
            print("  ".join(c.ljust(widths[c]) for c in cols))
            print("  ".join("─" * widths[c] for c in cols))
            for row in rows:
                print("  ".join(row[c].ljust(widths[c]) for c in cols))

    def _build_html(self, groups, sort_by, top) -> str:
        """Build a self-contained sortable HTML report."""
        sections = []
        for stype, objs in groups.items():
            cols = _COLUMNS.get(stype, _COLUMNS["_default"])
            color = {"indicator": "#1a73e8", "malware": "#d93025",
                     "vulnerability": "#f29900", "threat-actor": "#9334e6",
                     "relationship": "#0f9d58"}.get(stype, "#5f6368")
            rows_html = ""
            for obj in self._sort(objs, sort_by)[:top]:
                cells = "".join(
                    f"<td>{_coerce(_get_field(obj, c))[:200]}</td>"
                    for c in cols
                )
                rows_html += f"<tr>{cells}</tr>\n"
            headers = "".join(f"<th onclick='sortTable(this)'>{c} ⇅</th>" for c in cols)
            sections.append(f"""
    <div class="section">
      <h2 style="color:{color}">{stype} <span class="badge">{len(objs)}</span></h2>
      <input type="text" class="filter-input" placeholder="Filter {stype}..."
             onkeyup="filterTable(this)">
      <div class="table-wrap">
        <table class="stix-table" id="tbl-{stype}">
          <thead><tr>{headers}</tr></thead>
          <tbody>{rows_html}</tbody>
        </table>
      </div>
    </div>""")

        ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>GNAT Workspace: {self._ws.name}</title>
  <style>
    *{{box-sizing:border-box;margin:0;padding:0}}
    body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
         background:#0f1117;color:#e8eaed;padding:24px}}
    h1{{font-size:1.6rem;margin-bottom:4px;color:#8ab4f8}}
    .meta{{color:#9aa0a6;font-size:.85rem;margin-bottom:24px}}
    .badge{{background:#333;color:#e8eaed;border-radius:12px;
            padding:2px 10px;font-size:.8rem;margin-left:8px}}
    .section{{margin-bottom:40px}}
    h2{{font-size:1.1rem;margin-bottom:8px;font-weight:600}}
    .filter-input{{width:100%;max-width:420px;padding:6px 12px;margin-bottom:10px;
                   background:#1e2029;border:1px solid #3c4043;border-radius:6px;
                   color:#e8eaed;font-size:.9rem}}
    .table-wrap{{overflow-x:auto}}
    .stix-table{{width:100%;border-collapse:collapse;font-size:.82rem}}
    .stix-table thead tr{{background:#1e2029;position:sticky;top:0;z-index:1}}
    .stix-table th{{padding:8px 12px;text-align:left;cursor:pointer;
                    white-space:nowrap;border-bottom:2px solid #3c4043;
                    user-select:none}}
    .stix-table th:hover{{background:#2a2d3a}}
    .stix-table td{{padding:6px 12px;border-bottom:1px solid #2a2d3a;
                    max-width:360px;overflow:hidden;text-overflow:ellipsis;
                    white-space:nowrap}}
    .stix-table tr:hover td{{background:#1e2029}}
  </style>
</head>
<body>
  <h1>GNAT: {self._ws.name}</h1>
  <p class="meta">Generated {ts} &nbsp;|&nbsp; {len(self._ws)} total objects</p>
  {''.join(sections)}
  <script>
    function filterTable(input) {{
      const filter = input.value.toLowerCase();
      const table  = input.nextElementSibling.querySelector('table');
      for (const row of table.tBodies[0].rows) {{
        row.style.display = row.innerText.toLowerCase().includes(filter) ? '' : 'none';
      }}
    }}
    function sortTable(th) {{
      const table = th.closest('table');
      const col   = Array.from(th.parentNode.children).indexOf(th);
      const rows  = Array.from(table.tBodies[0].rows);
      const asc   = th.dataset.asc !== 'true';
      th.dataset.asc = asc;
      rows.sort((a, b) => {{
        const x = a.cells[col].innerText.trim();
        const y = b.cells[col].innerText.trim();
        const nx = parseFloat(x), ny = parseFloat(y);
        if (!isNaN(nx) && !isNaN(ny)) return asc ? nx - ny : ny - nx;
        return asc ? x.localeCompare(y) : y.localeCompare(x);
      }});
      rows.forEach(r => table.tBodies[0].appendChild(r));
    }}
  </script>
</body>
</html>"""
