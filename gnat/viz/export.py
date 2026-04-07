# SPDX-License-Identifier: Apache-2.0
# Copyright 2026 Bill Halpin
"""
gnat.viz.export
===================

Exports workspace data to external BI tools without requiring live API
connections.

Power BI
--------
Produces multi-sheet ``.xlsx`` files (via the existing
:meth:`~gnat.viz.tabular.TabularView.to_excel` method) with a column
structure designed for direct import into Power BI Desktop as a data source.
Also exports a Relationships table so Power BI can build relationship diagrams
in its own graph visual.

Grafana Dashboard JSON
-----------------------
Pre-built panel configurations for the most common GNAT use cases:
indicator risk heatmap, vulnerability CVSS timeline, enrichment activity,
type breakdown bar chart.  Copy the JSON into Grafana's dashboard import
dialog and point it at a running :class:`~gnat.viz.grafana.server.GrafanaServer`.
"""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from gnat.context.workspace import Workspace

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Power BI export
# ---------------------------------------------------------------------------


class PowerBIExporter:
    """
    Export workspace data in a Power BI-compatible format.

    Produces:

    * An ``.xlsx`` workbook with one sheet per STIX type and a separate
      ``Relationships`` sheet for the relationship graph visual.
    * An optional ``model.json`` describing the table schema and
      relationships for Power BI's data model.

    Parameters
    ----------
    workspace : Workspace
        Source workspace.

    Examples
    --------
    ::

        exporter = PowerBIExporter(workspace)
        exporter.to_xlsx("gnat_workspace.xlsx")
        exporter.to_model_json("model.json")
    """

    def __init__(self, workspace: Workspace):
        self._ws = workspace

    def to_xlsx(self, path: str) -> None:
        """
        Export all workspace objects to a Power BI-compatible Excel workbook.

        Sheet layout:
          * One sheet per STIX type (Indicator, Malware, Vulnerability, …)
          * ``Relationships`` sheet — all relationship objects as a flat table
          * ``EnrichmentLog`` sheet — full enrichment history
          * ``Summary`` sheet — object counts + workspace metadata

        Requires ``openpyxl``.
        """
        from gnat.viz.tabular import _COLUMNS, _coerce, _get_field

        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Power BI export: pip install 'gnat[viz]'")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill("solid", fgColor="1a73e8")
        header_font = Font(bold=True, color="FFFFFF")
        alt_fill = PatternFill("solid", fgColor="EEF4FF")

        def _write_sheet(sheet_name: str, columns: list[str], rows: list[list[Any]]) -> None:
            ws = wb.create_sheet(sheet_name[:31])
            ws.append(columns)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for i, row in enumerate(rows, start=2):
                ws.append(row)
                if i % 2 == 0:
                    for cell in ws[i]:
                        cell.fill = alt_fill
            ws.freeze_panes = "A2"
            for col_idx, col_name in enumerate(columns, start=1):
                col_letter = get_column_letter(col_idx)
                max_len = max((len(str(r[col_idx - 1])) for r in rows if r), default=0)
                ws.column_dimensions[col_letter].width = min(max(len(col_name), max_len) + 4, 60)

        # ── Object sheets ──────────────────────────────────────────────────
        by_type: dict[str, list] = {}
        for obj in self._ws.objects.values():
            by_type.setdefault(obj.stix_type, []).append(obj)

        for stype, objs in sorted(by_type.items()):
            if stype == "relationship":
                continue  # handled separately
            cols = _COLUMNS.get(stype, _COLUMNS["_default"])
            rows = [[_coerce(_get_field(obj, c)) for c in cols] for obj in objs]
            _write_sheet(stype.replace("-", " ").title(), cols, rows)

        # ── Relationships sheet (Power BI graph data model) ────────────────
        rel_cols = [
            "id",
            "relationship_type",
            "source_ref",
            "source_name",
            "target_ref",
            "target_name",
            "x_enrichment_source",
            "x_enrichment_strategy",
            "created",
        ]
        rel_rows = []
        for obj in self._ws.objects.values():
            if obj.stix_type != "relationship":
                continue
            src_id = obj._properties.get("source_ref", "")
            tgt_id = obj._properties.get("target_ref", "")
            src_obj = self._ws.objects.get(src_id)
            tgt_obj = self._ws.objects.get(tgt_id)
            src_name = getattr(src_obj, "name", src_id[:40]) if src_obj else src_id[:40]
            tgt_name = getattr(tgt_obj, "name", tgt_id[:40]) if tgt_obj else tgt_id[:40]
            rel_rows.append(
                [
                    obj.id,
                    obj._properties.get("relationship_type", ""),
                    src_id,
                    src_name,
                    tgt_id,
                    tgt_name,
                    obj._properties.get("x_enrichment_source", ""),
                    obj._properties.get("x_enrichment_strategy", ""),
                    obj._properties.get("created", ""),
                ]
            )
        _write_sheet("Relationships", rel_cols, rel_rows)

        # ── Enrichment log sheet ───────────────────────────────────────────
        enrich_hist = self._ws.get_enrichment_history()
        if enrich_hist:
            enr_cols = ["stix_id", "source_platform", "strategy", "created_at"]
            enr_rows = [
                [
                    e.get("stix_id", ""),
                    e.get("source_platform", ""),
                    e.get("strategy", ""),
                    e.get("created_at", ""),
                ]
                for e in enrich_hist
            ]
            _write_sheet("EnrichmentLog", enr_cols, enr_rows)

        # ── Summary sheet ──────────────────────────────────────────────────
        from datetime import datetime, timezone

        summary_ws = wb.create_sheet("Summary")
        summary_ws["A1"] = "GNAT Workspace Export"
        summary_ws["A1"].font = Font(bold=True, size=14)
        summary_data = [
            ("Workspace", self._ws.name),
            ("Description", getattr(self._ws, "description", "")),
            ("Exported", datetime.now(timezone.utc).isoformat()),
            ("Total Objects", len(self._ws)),
        ]
        for stype, objs in sorted(by_type.items()):
            summary_data.append((stype, len(objs)))
        for row_idx, (label, value) in enumerate(summary_data, start=2):
            summary_ws[f"A{row_idx}"] = label
            summary_ws[f"B{row_idx}"] = value
        summary_ws.column_dimensions["A"].width = 20
        summary_ws.column_dimensions["B"].width = 40

        wb.save(path)
        logger.info("PowerBIExporter: Excel written to %s", path)

    def to_model_json(self, path: str | None = None) -> dict:
        """
        Generate a Power BI data model descriptor JSON.

        Describes table names, column types, and relationships so Power BI
        Desktop can auto-configure the data model on import.

        Returns
        -------
        dict
            Model descriptor.  If *path* is provided, also writes to disk.
        """
        from gnat.viz.tabular import _COLUMNS

        by_type: dict[str, list] = {}
        for obj in self._ws.objects.values():
            by_type.setdefault(obj.stix_type, []).append(obj)

        tables = []
        for stype, _objs in sorted(by_type.items()):
            if stype == "relationship":
                continue
            cols = _COLUMNS.get(stype, _COLUMNS["_default"])
            tables.append(
                {
                    "name": stype.replace("-", " ").title(),
                    "columns": [
                        {
                            "name": c,
                            "dataType": "decimal"
                            if c in ("confidence", "x_cvss_score", "x_rf_risk_score")
                            else "text",
                        }
                        for c in cols
                    ],
                }
            )

        # Relationships table
        tables.append(
            {
                "name": "Relationships",
                "columns": [
                    {"name": "id", "dataType": "text"},
                    {"name": "relationship_type", "dataType": "text"},
                    {"name": "source_ref", "dataType": "text"},
                    {"name": "source_name", "dataType": "text"},
                    {"name": "target_ref", "dataType": "text"},
                    {"name": "target_name", "dataType": "text"},
                    {"name": "x_enrichment_source", "dataType": "text"},
                    {"name": "x_enrichment_strategy", "dataType": "text"},
                    {"name": "created", "dataType": "dateTime"},
                ],
            }
        )

        model = {
            "name": f"GNAT_{self._ws.name}",
            "tables": tables,
            "relationships": [
                {
                    "name": "Relationships_to_Source",
                    "fromTable": "Relationships",
                    "fromColumn": "source_ref",
                    "toTable": t["name"],
                    "toColumn": "id",
                    "crossFilteringBehavior": "oneDirection",
                }
                for t in tables
                if t["name"] != "Relationships"
            ],
        }
        if path:
            Path(path).write_text(json.dumps(model, indent=2), encoding="utf-8")
            logger.info("PowerBIExporter: model JSON written to %s", path)
        return model


# ---------------------------------------------------------------------------
# Pre-built Grafana dashboard templates
# ---------------------------------------------------------------------------


def grafana_dashboard(
    workspace_name: str,
    datasource_name: str = "GNAT",
    title: str | None = None,
) -> dict:
    """
    Generate a pre-built Grafana dashboard JSON for a GNAT workspace.

    Includes panels for:

    * **Object Type Breakdown** — bar chart of STIX object counts by type
    * **Indicator Risk Score Timeline** — time-series of RF risk scores
    * **Vulnerability CVSS Heatmap** — CVSS scores over time
    * **Enrichment Activity** — annotation-driven enrichment events
    * **Indicator Table** — live table of all indicators
    * **Relationship Summary** — table of all relationship objects

    Parameters
    ----------
    workspace_name : str
        The workspace to target in the datasource queries.
    datasource_name : str
        The Grafana datasource name for the SimpleJSON source.
    title : str, optional
        Dashboard title.  Defaults to ``"GNAT: <workspace_name>"``.

    Returns
    -------
    dict
        Grafana dashboard JSON.  Import via Grafana → Dashboards → Import → JSON.

    Examples
    --------
    ::

        dashboard = grafana_dashboard("apt28-investigation")
        with open("grafana_dashboard.json", "w") as f:
            json.dump(dashboard, f, indent=2)
    """
    ds = {"type": "simplejson", "uid": datasource_name}
    title = title or f"GNAT: {workspace_name}"

    def _target(target_str, ref_id="A"):
        return {"target": target_str, "refId": ref_id, "type": "timeserie"}

    def _table_target(target_str, ref_id="A"):
        return {"target": target_str, "refId": ref_id, "type": "table"}

    panels = [
        # ── 1. Object count by type ────────────────────────────────────────
        {
            "id": 1,
            "title": "Object Count by Type",
            "type": "barchart",
            "gridPos": {"x": 0, "y": 0, "w": 8, "h": 8},
            "datasource": ds,
            "targets": [_table_target(f"{workspace_name}/summary")],
            "options": {
                "barWidth": 0.9,
                "fillOpacity": 80,
                "gradientMode": "opacity",
            },
        },
        # ── 2. Indicator Risk Score timeline ──────────────────────────────
        {
            "id": 2,
            "title": "Indicator RF Risk Scores",
            "type": "timeseries",
            "gridPos": {"x": 8, "y": 0, "w": 16, "h": 8},
            "datasource": ds,
            "targets": [_target(f"{workspace_name}/indicator/x_rf_risk_score")],
            "fieldConfig": {
                "defaults": {
                    "color": {"mode": "thresholds"},
                    "thresholds": {
                        "steps": [
                            {"color": "green", "value": None},
                            {"color": "yellow", "value": 25},
                            {"color": "orange", "value": 65},
                            {"color": "red", "value": 90},
                        ]
                    },
                    "min": 0,
                    "max": 100,
                }
            },
        },
        # ── 3. Vulnerability CVSS scores ──────────────────────────────────
        {
            "id": 3,
            "title": "Vulnerability CVSS Scores",
            "type": "timeseries",
            "gridPos": {"x": 0, "y": 8, "w": 12, "h": 8},
            "datasource": ds,
            "targets": [_target(f"{workspace_name}/vulnerability/x_cvss_score")],
            "fieldConfig": {
                "defaults": {
                    "color": {"mode": "thresholds"},
                    "thresholds": {
                        "steps": [
                            {"color": "green", "value": None},
                            {"color": "yellow", "value": 4.0},
                            {"color": "orange", "value": 7.0},
                            {"color": "red", "value": 9.0},
                        ]
                    },
                    "min": 0,
                    "max": 10,
                }
            },
        },
        # ── 4. Indicator confidence ────────────────────────────────────────
        {
            "id": 4,
            "title": "Indicator Confidence",
            "type": "gauge",
            "gridPos": {"x": 12, "y": 8, "w": 4, "h": 8},
            "datasource": ds,
            "targets": [_target(f"{workspace_name}/indicator/confidence")],
            "options": {"reduceOptions": {"calcs": ["mean"]}},
            "fieldConfig": {
                "defaults": {
                    "min": 0,
                    "max": 100,
                    "thresholds": {
                        "steps": [
                            {"color": "red", "value": None},
                            {"color": "yellow", "value": 40},
                            {"color": "green", "value": 70},
                        ]
                    },
                }
            },
        },
        # ── 5. Indicator table ─────────────────────────────────────────────
        {
            "id": 5,
            "title": "Indicators",
            "type": "table",
            "gridPos": {"x": 0, "y": 16, "w": 24, "h": 10},
            "datasource": ds,
            "targets": [_table_target(f"{workspace_name}/indicator")],
            "options": {
                "sortBy": [{"displayName": "confidence", "desc": True}],
                "footer": {"show": True, "reducer": ["count"], "fields": ["name"]},
            },
        },
        # ── 6. Relationships table ─────────────────────────────────────────
        {
            "id": 6,
            "title": "Relationships",
            "type": "table",
            "gridPos": {"x": 0, "y": 26, "w": 24, "h": 8},
            "datasource": ds,
            "targets": [_table_target(f"{workspace_name}/relationship")],
        },
    ]

    return {
        "__inputs": [
            {
                "name": datasource_name,
                "label": "GNAT Datasource",
                "description": "SimpleJSON datasource pointing at gnat viz serve",
                "type": "datasource",
                "pluginId": "simplejson",
                "pluginName": "SimpleJSON",
            }
        ],
        "annotations": {
            "list": [
                {
                    "datasource": ds,
                    "enable": True,
                    "name": "Enrichment Events",
                    "query": workspace_name,
                    "iconColor": "blue",
                }
            ]
        },
        "title": title,
        "uid": f"ctmsak-{workspace_name}",
        "schemaVersion": 38,
        "version": 1,
        "refresh": "30s",
        "time": {"from": "now-30d", "to": "now"},
        "panels": panels,
        "tags": ["gnat", "threat-intelligence", workspace_name],
    }


def save_grafana_dashboard(
    workspace_name: str,
    path: str,
    datasource_name: str = "GNAT",
    title: str | None = None,
) -> None:
    """
    Write a pre-built Grafana dashboard JSON file to disk.

    Parameters
    ----------
    workspace_name : str
        Workspace to target.
    path : str
        Output ``.json`` file path.
    """
    dashboard = grafana_dashboard(workspace_name, datasource_name, title)
    Path(path).write_text(json.dumps(dashboard, indent=2), encoding="utf-8")
    logger.info("Grafana dashboard JSON written to %s", path)
    print(f"Dashboard saved: {path}")
    print("To import: Grafana → Dashboards → Import → Upload JSON file")


def solr_dashboard(
    datasource_name: str = "GNAT-Solr",
    title: str = "GNAT Search Index",
) -> dict:
    """
    Generate a pre-built Grafana dashboard JSON for the GNAT Solr search sidecar.

    Targets the ``/solr/`` sub-router on the GNAT Grafana datasource server.

    Panels
    ------
    * **Total Indexed Documents** — single stat
    * **Objects by STIX Type** — bar chart
    * **Objects by Source Platform** — bar chart
    * **Ingest Rate (docs/day)** — time-series
    * **Full Index Breakdown (type × platform)** — table with both facets
    * **Recent Ingest Activity** — table of type counts

    Parameters
    ----------
    datasource_name : str
        The Grafana datasource name pointing at the GNAT server's ``/solr/``
        base URL.  Default ``"GNAT-Solr"``.
    title : str
        Dashboard title.  Default ``"GNAT Search Index"``.

    Returns
    -------
    dict
        Grafana dashboard JSON ready for import.

    Examples
    --------
    ::

        from gnat.viz.export import solr_dashboard
        import json, pathlib
        pathlib.Path("solr_dashboard.json").write_text(
            json.dumps(solr_dashboard(), indent=2)
        )
    """
    ds = {"type": "simplejson", "uid": datasource_name}

    def _tgt(target_str: str, ref_id: str = "A") -> dict:
        return {"target": target_str, "refId": ref_id, "type": "timeserie"}

    def _tbl(target_str: str, ref_id: str = "A") -> dict:
        return {"target": target_str, "refId": ref_id, "type": "table"}

    panels = [
        # ── 1. Total indexed documents (stat) ─────────────────────────────
        {
            "id": 1,
            "title": "Total Indexed Documents",
            "type": "stat",
            "gridPos": {"h": 4, "w": 4, "x": 0, "y": 0},
            "datasource": ds,
            "targets": [_tbl("stats/total")],
            "options": {
                "reduceOptions": {"calcs": ["lastNotNull"]},
                "colorMode": "value",
                "graphMode": "none",
            },
            "fieldConfig": {
                "defaults": {
                    "color": {"mode": "thresholds"},
                    "thresholds": {
                        "steps": [
                            {"color": "green", "value": 0},
                            {"color": "yellow", "value": 1000},
                            {"color": "red", "value": 100000},
                        ]
                    },
                },
            },
        },
        # ── 2. Objects by STIX Type (bar chart) ──────────────────────────
        {
            "id": 2,
            "title": "Objects by STIX Type",
            "type": "barchart",
            "gridPos": {"h": 8, "w": 10, "x": 4, "y": 0},
            "datasource": ds,
            "targets": [_tbl("facet/stix_type")],
            "fieldConfig": {
                "defaults": {"color": {"mode": "palette-classic"}},
            },
            "options": {"xField": "STIX Type"},
        },
        # ── 3. Objects by Source Platform (bar chart) ─────────────────────
        {
            "id": 3,
            "title": "Objects by Source Platform",
            "type": "barchart",
            "gridPos": {"h": 8, "w": 10, "x": 14, "y": 0},
            "datasource": ds,
            "targets": [_tbl("facet/source_platform")],
            "fieldConfig": {
                "defaults": {"color": {"mode": "palette-classic"}},
            },
            "options": {"xField": "source_platform"},
        },
        # ── 4. Ingest rate (time-series docs/day) ─────────────────────────
        {
            "id": 4,
            "title": "Ingest Rate (documents / day)",
            "type": "timeseries",
            "gridPos": {"h": 8, "w": 12, "x": 0, "y": 8},
            "datasource": ds,
            "targets": [_tgt("timeseries/ingest")],
            "fieldConfig": {
                "defaults": {
                    "color": {"mode": "palette-classic"},
                    "custom": {"lineWidth": 2, "fillOpacity": 10},
                },
            },
            "options": {"tooltip": {"mode": "single"}},
        },
        # ── 5. Type × platform breakdown (table) ─────────────────────────
        {
            "id": 5,
            "title": "Type Breakdown",
            "type": "table",
            "gridPos": {"h": 8, "w": 12, "x": 12, "y": 8},
            "datasource": ds,
            "targets": [_tbl("stats/type_counts", "A"), _tbl("stats/platform_counts", "B")],
            "options": {"sortBy": [{"displayName": "Doc Count", "desc": True}]},
        },
        # ── 6. Platform counts table ──────────────────────────────────────
        {
            "id": 6,
            "title": "Platform Counts",
            "type": "table",
            "gridPos": {"h": 6, "w": 12, "x": 0, "y": 16},
            "datasource": ds,
            "targets": [_tbl("stats/platform_counts")],
            "options": {"sortBy": [{"displayName": "Doc Count", "desc": True}]},
        },
        # ── 7. Search result table (variable-driven) ───────────────────────
        {
            "id": 7,
            "title": "Search Results",
            "type": "table",
            "gridPos": {"h": 6, "w": 12, "x": 12, "y": 16},
            "datasource": ds,
            "targets": [_tbl("search/*:*")],
            "description": (
                "Change target to 'search/<your-query>' to filter by any indexed text."
            ),
        },
    ]

    return {
        "uid": "gnat-solr-index",
        "title": title,
        "schemaVersion": 38,
        "version": 1,
        "refresh": "1m",
        "time": {"from": "now-30d", "to": "now"},
        "panels": panels,
        "tags": ["gnat", "solr", "search", "threat-intelligence"],
        "templating": {"list": []},
        "annotations": {"list": []},
    }


def save_solr_dashboard(
    path: str,
    datasource_name: str = "GNAT-Solr",
    title: str = "GNAT Search Index",
) -> None:
    """
    Write the Solr search sidecar Grafana dashboard JSON to disk.

    Parameters
    ----------
    path : str
        Output ``.json`` file path.
    datasource_name : str
        Grafana datasource name for the ``/solr/`` endpoints.
    title : str
        Dashboard title.
    """
    dashboard = solr_dashboard(datasource_name=datasource_name, title=title)
    Path(path).write_text(json.dumps(dashboard, indent=2), encoding="utf-8")
    logger.info("Solr dashboard JSON written to %s", path)
    print(f"Solr dashboard saved: {path}")
    print("To import: Grafana → Dashboards → Import → Upload JSON file")
    print("Configure datasource: SimpleJSON → http://localhost:3001")
