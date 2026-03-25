"""
tests/unit/viz/test_viz.py
===========================

Unit tests for ctm_sak.viz — tabular, graph, export, Grafana server.

No real network calls, no Plotly rendering, no file system side-effects
beyond tmp_path fixtures.  All external deps (plotly, networkx, openpyxl,
fastapi) are tested with graceful skips if not installed.
"""

from __future__ import annotations

import json
import tempfile
from pathlib import Path
from typing import Iterator
from unittest.mock import MagicMock, patch

import pytest

from ctm_sak.context import FlatFileStore, GlobalContext, GlobalContextRegistry, Workspace
from ctm_sak.orm.indicator import Indicator
from ctm_sak.orm.malware import Malware
from ctm_sak.orm.vulnerability import Vulnerability
from ctm_sak.orm.relationship import Relationship
from ctm_sak.viz.tabular import TabularView, _coerce, _get_field, _to_rows
from ctm_sak.viz.graph import GraphView
from ctm_sak.viz.export import PowerBIExporter, grafana_dashboard, save_grafana_dashboard


# ===========================================================================
# Fixtures
# ===========================================================================

def _mock_gc(name="threatq"):
    cli = MagicMock()
    cli.target = name
    cli.ping.return_value = True
    cli.client = MagicMock()
    cli.client.list_objects.return_value = []
    cli.client.to_stix.return_value = {}
    return GlobalContext(name=name, client=cli)


def _make_registry():
    r = GlobalContextRegistry()
    r.register(_mock_gc())
    r.set_default("threatq")
    return r


def _populated_workspace(tmp_path, n_indicators=5, n_malware=2,
                          n_vulns=3, n_rels=4) -> Workspace:
    store    = FlatFileStore(base_dir=str(tmp_path / "workspaces"))
    registry = _make_registry()
    ws       = Workspace("test", registry, store)

    indicators = []
    for i in range(n_indicators):
        ind = Indicator(
            name=f"evil-{i}.com",
            pattern=f"[domain-name:value = 'evil-{i}.com']",
            pattern_type="stix",
            indicator_types=["malicious-activity"],
            confidence=50 + i * 8,
            x_tlp="amber",
            x_rf_risk_score=30 + i * 12,
            created=f"2024-01-{i+1:02d}T00:00:00Z",
            modified=f"2024-01-{i+1:02d}T00:00:00Z",
        )
        ws.add(ind, mark_dirty=False)
        indicators.append(ind)

    malwares = []
    for i in range(n_malware):
        mal = Malware(name=f"BadMal-{i}", is_family=False,
                      confidence=70, created=f"2024-02-0{i+1}T00:00:00Z",
                      modified=f"2024-02-0{i+1}T00:00:00Z")
        ws.add(mal, mark_dirty=False)
        malwares.append(mal)

    vulns = []
    for i in range(n_vulns):
        v = Vulnerability(name=f"CVE-2024-{i:04d}",
                          x_cvss_score=5.0 + i,
                          confidence=80,
                          created=f"2024-03-0{i+1}T00:00:00Z",
                          modified=f"2024-03-0{i+1}T00:00:00Z")
        ws.add(v, mark_dirty=False)
        vulns.append(v)

    # Add relationships between indicators and malware
    for i in range(min(n_rels, len(indicators), len(malwares))):
        rel = Relationship(
            relationship_type="indicates",
            source_ref=indicators[i % len(indicators)].id,
            target_ref=malwares[i % len(malwares)].id,
            x_enrichment_source="recorded_future",
            x_enrichment_strategy="create_relationships",
            created=f"2024-04-0{i+1}T00:00:00Z",
            modified=f"2024-04-0{i+1}T00:00:00Z",
        )
        ws.add(rel, mark_dirty=False)

    return ws


# ===========================================================================
# TabularView — helper functions
# ===========================================================================

class TestTabularHelpers:

    def test_coerce_none(self):
        assert _coerce(None) == ""

    def test_coerce_list(self):
        assert _coerce(["a", "b"]) == "a, b"

    def test_coerce_bool_true(self):
        assert _coerce(True) == "yes"

    def test_coerce_bool_false(self):
        assert _coerce(False) == "no"

    def test_coerce_float(self):
        assert _coerce(9.84567) == "9.8"

    def test_coerce_string(self):
        assert _coerce("hello") == "hello"

    def test_get_field_stix_type(self):
        ind = Indicator(name="test")
        assert _get_field(ind, "type") == "indicator"

    def test_get_field_property(self):
        ind = Indicator(name="test.com")
        assert _get_field(ind, "name") == "test.com"

    def test_get_field_missing(self):
        ind = Indicator()
        assert _get_field(ind, "nonexistent_field") is None

    def test_get_field_x_property(self):
        ind = Indicator(x_rf_risk_score=90)
        assert _get_field(ind, "x_rf_risk_score") == 90

    def test_to_rows(self):
        inds = [Indicator(name="a.com", confidence=80),
                Indicator(name="b.com", confidence=60)]
        rows = _to_rows(inds, ["name", "confidence"])
        assert rows[0]["name"] == "a.com"
        assert rows[0]["confidence"] == "80"
        assert len(rows) == 2


# ===========================================================================
# TabularView — terminal output
# ===========================================================================

class TestTabularView:

    def test_show_plain_no_crash(self, tmp_path, capsys):
        ws   = _populated_workspace(tmp_path)
        view = TabularView(ws)
        # Patch rich to force plain fallback
        with patch("ctm_sak.viz.tabular.TabularView._show_rich",
                   side_effect=ImportError("no rich")):
            view.show()
        captured = capsys.readouterr()
        assert "indicator" in captured.out.lower()

    def test_show_filters_by_type(self, tmp_path, capsys):
        ws   = _populated_workspace(tmp_path)
        view = TabularView(ws)
        with patch("ctm_sak.viz.tabular.TabularView._show_rich",
                   side_effect=ImportError("no rich")):
            view.show(stix_type="malware")
        captured = capsys.readouterr()
        assert "malware" in captured.out.lower()

    def test_group_objects_all_types(self, tmp_path):
        ws    = _populated_workspace(tmp_path)
        view  = TabularView(ws)
        groups = view._group_objects()
        assert "indicator" in groups
        assert "malware" in groups
        assert "vulnerability" in groups
        assert "relationship" in groups

    def test_group_objects_filtered(self, tmp_path):
        ws    = _populated_workspace(tmp_path)
        view  = TabularView(ws)
        groups = view._group_objects("indicator")
        assert list(groups.keys()) == ["indicator"]

    def test_sort_by_confidence_descending(self, tmp_path):
        ws   = _populated_workspace(tmp_path, n_indicators=4, n_malware=0,
                                    n_vulns=0, n_rels=0)
        view = TabularView(ws)
        inds = list(ws.objects.values())
        sorted_inds = view._sort(inds, "confidence")
        confs = [o._properties.get("confidence", 0) for o in sorted_inds]
        assert confs == sorted(confs, reverse=True)


# ===========================================================================
# TabularView — HTML export
# ===========================================================================

class TestTabularHtml:

    def test_to_html_returns_string(self, tmp_path):
        ws   = _populated_workspace(tmp_path)
        view = TabularView(ws)
        html = view.to_html()
        assert isinstance(html, str)
        assert "<!DOCTYPE html>" in html

    def test_to_html_contains_type_names(self, tmp_path):
        ws   = _populated_workspace(tmp_path)
        view = TabularView(ws)
        html = view.to_html()
        assert "indicator" in html
        assert "malware" in html

    def test_to_html_contains_object_names(self, tmp_path):
        ws   = _populated_workspace(tmp_path)
        view = TabularView(ws)
        html = view.to_html()
        assert "evil-0.com" in html

    def test_to_html_writes_file(self, tmp_path):
        ws   = _populated_workspace(tmp_path)
        view = TabularView(ws)
        out  = str(tmp_path / "report.html")
        view.to_html(out)
        assert Path(out).exists()
        assert Path(out).stat().st_size > 1000

    def test_to_html_sortable_js_present(self, tmp_path):
        ws   = _populated_workspace(tmp_path)
        html = TabularView(ws).to_html()
        assert "sortTable" in html
        assert "filterTable" in html

    def test_to_html_dark_theme(self, tmp_path):
        ws   = _populated_workspace(tmp_path)
        html = TabularView(ws).to_html()
        assert "#0f1117" in html  # dark background


# ===========================================================================
# TabularView — CSV export
# ===========================================================================

class TestTabularCsv:

    def test_to_csv_creates_file(self, tmp_path):
        ws  = _populated_workspace(tmp_path)
        out = str(tmp_path / "output.csv")
        TabularView(ws).to_csv(out)
        assert Path(out).exists()

    def test_to_csv_has_header(self, tmp_path):
        ws  = _populated_workspace(tmp_path)
        out = str(tmp_path / "output.csv")
        TabularView(ws).to_csv(out)
        content = Path(out).read_text()
        assert "name" in content.splitlines()[0].lower()

    def test_to_csv_has_data_rows(self, tmp_path):
        ws  = _populated_workspace(tmp_path, n_indicators=3)
        out = str(tmp_path / "output.csv")
        TabularView(ws).to_csv(out)
        lines = Path(out).read_text().splitlines()
        assert len(lines) > 1  # header + at least one data row

    def test_to_csv_filtered_by_type(self, tmp_path):
        ws  = _populated_workspace(tmp_path)
        out = str(tmp_path / "output.csv")
        TabularView(ws).to_csv(out, stix_type="indicator")
        content = Path(out).read_text()
        assert "evil" in content


# ===========================================================================
# TabularView — Excel export
# ===========================================================================

class TestTabularExcel:

    def test_to_excel_requires_openpyxl(self, tmp_path):
        ws  = _populated_workspace(tmp_path)
        out = str(tmp_path / "output.xlsx")
        with patch.dict("sys.modules", {"openpyxl": None}):
            with pytest.raises(ImportError, match="openpyxl"):
                TabularView(ws).to_excel(out)

    @pytest.mark.skipif(
        not __import__("importlib").util.find_spec("openpyxl"),
        reason="openpyxl not installed"
    )
    def test_to_excel_creates_file(self, tmp_path):
        ws  = _populated_workspace(tmp_path)
        out = str(tmp_path / "output.xlsx")
        TabularView(ws).to_excel(out)
        assert Path(out).exists()
        assert Path(out).stat().st_size > 1000

    @pytest.mark.skipif(
        not __import__("importlib").util.find_spec("openpyxl"),
        reason="openpyxl not installed"
    )
    def test_to_excel_has_multiple_sheets(self, tmp_path):
        import openpyxl
        ws  = _populated_workspace(tmp_path)
        out = str(tmp_path / "output.xlsx")
        TabularView(ws).to_excel(out)
        wb = openpyxl.load_workbook(out)
        assert "Summary" in wb.sheetnames
        assert any("Indicator" in s for s in wb.sheetnames)


# ===========================================================================
# GraphView — structure
# ===========================================================================

class TestGraphView:

    # ── Graph extraction ─────────────────────────────────────────────────

    def test_extract_graph_nodes_and_edges(self, tmp_path):
        ws   = _populated_workspace(tmp_path)
        gv   = GraphView(ws)
        nodes, edges = gv._extract_graph()
        assert len(nodes) > 0
        assert len(edges) > 0

    def test_extract_graph_filters_by_type_contains_target(self, tmp_path):
        ws   = _populated_workspace(tmp_path)
        gv   = GraphView(ws)
        nodes, _ = gv._extract_graph(stix_types=["indicator"])
        assert any(obj.stix_type == "indicator" for obj in nodes.values())

    def test_extract_graph_no_rels(self, tmp_path):
        ws = _populated_workspace(tmp_path, n_rels=0)
        _, edges = GraphView(ws)._extract_graph()
        assert edges == []

    def test_extract_graph_relationship_filter(self, tmp_path):
        ws = _populated_workspace(tmp_path)
        _, edges = GraphView(ws)._extract_graph(relationship_types=["uses"])
        assert edges == []

    def test_top_by_degree_caps_correctly(self, tmp_path):
        ws     = _populated_workspace(tmp_path)
        gv     = GraphView(ws)
        nodes, edges = gv._extract_graph()
        capped = gv._top_by_degree(nodes, edges, 5)
        assert len(capped) == 5

    def test_max_nodes_parameter(self, tmp_path):
        ws = _populated_workspace(tmp_path, n_indicators=20, n_malware=5,
                                   n_vulns=5, n_rels=10)
        nodes, _ = GraphView(ws)._extract_graph(max_nodes=10)
        assert len(nodes) <= 10

    # ── Layout algorithms ─────────────────────────────────────────────────

    def test_barnes_hut_layout_covers_all_nodes(self, tmp_path):
        import random as _r, math as _m
        from ctm_sak.viz.graph import _barnes_hut_layout
        n   = 100
        ids = [f"n{i}" for i in range(n)]
        rng = _r.Random(42)
        adj = {nid: [ids[rng.randint(0, n-1)]] for nid in ids}
        pos = _barnes_hut_layout(ids, adj, iterations=20, seed=42)
        assert len(pos) == n
        for x, y in pos.values():
            assert _m.isfinite(x) and _m.isfinite(y)

    def test_type_cluster_layout(self, tmp_path):
        import random as _r
        from ctm_sak.viz.graph import _type_cluster_layout
        n    = 200
        ids  = [f"n{i}" for i in range(n)]
        rng  = _r.Random(42)
        types = {nid: rng.choice(["indicator", "malware"]) for nid in ids}
        pos  = _type_cluster_layout(ids, types, {}, seed=42)
        assert len(pos) == n

    def test_compute_layout_returns_all_nodes(self, tmp_path):
        ws = _populated_workspace(tmp_path, n_indicators=5, n_malware=2,
                                   n_vulns=2, n_rels=4)
        gv = GraphView(ws)
        nodes, edges = gv._extract_graph()
        pos = gv._compute_layout(nodes, edges)
        assert set(pos.keys()) == set(nodes.keys())
        import math
        for x, y in pos.values():
            assert math.isfinite(x) and math.isfinite(y)

    def test_compute_layout_cluster_path(self, tmp_path):
        ws = _populated_workspace(tmp_path, n_indicators=30, n_malware=10,
                                   n_vulns=10, n_rels=20)
        gv = GraphView(ws, cluster_threshold=10)
        nodes, edges = gv._extract_graph()
        pos = gv._compute_layout(nodes, edges)
        assert len(pos) == len(nodes)

    # ── Node sizing ───────────────────────────────────────────────────────

    def test_node_size_scales_with_confidence(self, tmp_path):
        gv   = GraphView(_populated_workspace(tmp_path))
        assert gv._node_size(Indicator(confidence=0)) < gv._node_size(Indicator(confidence=100))

    def test_node_size_fallback(self, tmp_path):
        size = GraphView(_populated_workspace(tmp_path))._node_size(Indicator())
        assert 4 <= size <= 22

    # ── Output methods ────────────────────────────────────────────────────

    def test_to_graph_json_structure(self, tmp_path):
        data = GraphView(_populated_workspace(tmp_path)).to_graph_json()
        assert "nodes" in data and "edges" in data
        for node in data["nodes"]:
            assert "key" in node and "x" in node and "y" in node

    def test_to_graph_json_writes_file(self, tmp_path):
        out = str(tmp_path / "graph.json")
        GraphView(_populated_workspace(tmp_path)).to_graph_json(out)
        assert Path(out).exists()
        assert "nodes" in json.loads(Path(out).read_text())

    def test_to_html_sigma_creates_file(self, tmp_path):
        out = str(tmp_path / "graph.html")
        GraphView(_populated_workspace(tmp_path)).to_html(out, renderer="sigma")
        assert Path(out).exists()
        content = Path(out).read_text()
        assert "GRAPH_DATA" in content
        assert "indicator" in content

    def test_to_html_sigma_has_controls(self, tmp_path):
        out = str(tmp_path / "graph2.html")
        GraphView(_populated_workspace(tmp_path)).to_html(out, renderer="sigma")
        content = Path(out).read_text()
        assert "resetCamera" in content
        assert "filterByType" in content
        assert "type-filter" in content

    def test_to_html_plotly_creates_file(self, tmp_path):
        ws  = _populated_workspace(tmp_path, n_indicators=5, n_malware=2,
                                    n_vulns=2, n_rels=4)
        out = str(tmp_path / "graph3.html")
        try:
            GraphView(ws).to_html(out, renderer="plotly3d")
            assert Path(out).exists()
        except ImportError:
            pytest.skip("plotly not installed")

    def test_to_json_returns_plotly_spec(self, tmp_path):
        try:
            spec = GraphView(_populated_workspace(tmp_path)).to_json()
            d    = json.loads(spec)
            assert "data" in d or "layout" in d
        except ImportError:
            pytest.skip("plotly not installed")

    def test_to_networkx(self, tmp_path):
        try:
            G = GraphView(_populated_workspace(tmp_path)).to_networkx()
            assert G.number_of_nodes() > 0
        except ImportError:
            pytest.skip("networkx not installed")

    def test_summary(self, tmp_path):
        s = GraphView(_populated_workspace(tmp_path)).summary()
        assert "nodes" in s and "edges" in s
        assert "indicator" in s["node_types"]





# ===========================================================================
# PowerBIExporter
# ===========================================================================

class TestPowerBIExporter:

    def test_to_model_json_structure(self, tmp_path):
        ws   = _populated_workspace(tmp_path)
        exp  = PowerBIExporter(ws)
        model = exp.to_model_json()
        assert "name" in model
        assert "tables" in model
        assert "relationships" in model
        table_names = [t["name"] for t in model["tables"]]
        assert "Relationships" in table_names

    def test_to_model_json_has_columns(self, tmp_path):
        ws    = _populated_workspace(tmp_path)
        model = PowerBIExporter(ws).to_model_json()
        for table in model["tables"]:
            assert "columns" in table
            assert len(table["columns"]) > 0

    def test_to_model_json_writes_file(self, tmp_path):
        ws  = _populated_workspace(tmp_path)
        out = str(tmp_path / "model.json")
        PowerBIExporter(ws).to_model_json(out)
        assert Path(out).exists()
        loaded = json.loads(Path(out).read_text())
        assert "tables" in loaded

    def test_to_model_json_relationships_reference_tables(self, tmp_path):
        ws    = _populated_workspace(tmp_path)
        model = PowerBIExporter(ws).to_model_json()
        table_names = {t["name"] for t in model["tables"]}
        for rel in model["relationships"]:
            assert rel["fromTable"] in table_names or rel["toTable"] in table_names

    @pytest.mark.skipif(
        not __import__("importlib").util.find_spec("openpyxl"),
        reason="openpyxl not installed"
    )
    def test_to_xlsx_creates_file(self, tmp_path):
        ws  = _populated_workspace(tmp_path)
        out = str(tmp_path / "powerbi.xlsx")
        PowerBIExporter(ws).to_xlsx(out)
        assert Path(out).exists()

    @pytest.mark.skipif(
        not __import__("importlib").util.find_spec("openpyxl"),
        reason="openpyxl not installed"
    )
    def test_to_xlsx_has_relationships_sheet(self, tmp_path):
        import openpyxl
        ws  = _populated_workspace(tmp_path)
        out = str(tmp_path / "powerbi.xlsx")
        PowerBIExporter(ws).to_xlsx(out)
        wb = openpyxl.load_workbook(out)
        assert "Relationships" in wb.sheetnames

    @pytest.mark.skipif(
        not __import__("importlib").util.find_spec("openpyxl"),
        reason="openpyxl not installed"
    )
    def test_to_xlsx_has_summary_sheet(self, tmp_path):
        import openpyxl
        ws  = _populated_workspace(tmp_path)
        out = str(tmp_path / "powerbi.xlsx")
        PowerBIExporter(ws).to_xlsx(out)
        wb = openpyxl.load_workbook(out)
        assert "Summary" in wb.sheetnames


# ===========================================================================
# Grafana dashboard JSON
# ===========================================================================

class TestGrafanaDashboard:

    def test_grafana_dashboard_structure(self):
        d = grafana_dashboard("apt28")
        assert d["title"] == "CTM-SAK: apt28"
        assert d["schemaVersion"] == 38
        assert "panels" in d
        assert len(d["panels"]) >= 4

    def test_grafana_dashboard_custom_title(self):
        d = grafana_dashboard("apt28", title="My Custom Dashboard")
        assert d["title"] == "My Custom Dashboard"

    def test_grafana_dashboard_uid(self):
        d = grafana_dashboard("apt28")
        assert "apt28" in d["uid"]

    def test_grafana_dashboard_panels_have_datasource(self):
        d = grafana_dashboard("apt28", datasource_name="MyDS")
        for panel in d["panels"]:
            assert panel["datasource"]["uid"] == "MyDS"

    def test_grafana_dashboard_targets_reference_workspace(self):
        d = grafana_dashboard("my-ws")
        for panel in d["panels"]:
            for target in panel.get("targets", []):
                assert "my-ws" in target.get("target", "")

    def test_save_grafana_dashboard(self, tmp_path):
        out = str(tmp_path / "dashboard.json")
        save_grafana_dashboard("apt28", out)
        assert Path(out).exists()
        loaded = json.loads(Path(out).read_text())
        assert loaded["title"] == "CTM-SAK: apt28"

    def test_grafana_dashboard_has_annotations(self):
        d = grafana_dashboard("apt28")
        assert "annotations" in d
        assert len(d["annotations"]["list"]) > 0

    def test_grafana_dashboard_tags(self):
        d = grafana_dashboard("apt28")
        assert "threat-intelligence" in d["tags"]
        assert "apt28" in d["tags"]


# ===========================================================================
# Grafana server (FastAPI) — requires fastapi
# ===========================================================================

class TestGrafanaServer:

    @pytest.fixture
    def manager(self, tmp_path):
        from ctm_sak.context import WorkspaceManager
        store    = FlatFileStore(base_dir=str(tmp_path / "workspaces"))
        registry = _make_registry()
        mgr      = WorkspaceManager(registry, store=store)
        ws       = mgr.create("test-ws")
        ws.add(Indicator(name="evil.com",
                         pattern="[domain-name:value = 'evil.com']",
                         pattern_type="stix",
                         confidence=80,
                         created="2024-01-01T00:00:00Z",
                         modified="2024-01-01T00:00:00Z"),
               mark_dirty=False)
        return mgr

    def test_build_app_requires_fastapi(self, manager):
        try:
            from ctm_sak.viz.grafana.server import build_app
            app = build_app(manager)
            assert app is not None
        except ImportError:
            pytest.skip("fastapi not installed")

    def test_health_endpoint(self, manager):
        try:
            from fastapi.testclient import TestClient
            from ctm_sak.viz.grafana.server import build_app
            client = TestClient(build_app(manager))
            resp   = client.get("/")
            assert resp.status_code == 200
            assert resp.json()["status"] == "ok"
        except ImportError:
            pytest.skip("fastapi/httpx not installed")

    def test_workspaces_endpoint(self, manager):
        try:
            from fastapi.testclient import TestClient
            from ctm_sak.viz.grafana.server import build_app
            client = TestClient(build_app(manager))
            resp   = client.get("/workspaces")
            assert resp.status_code == 200
            names  = [w["name"] for w in resp.json()]
            assert "test-ws" in names
        except ImportError:
            pytest.skip("fastapi not installed")

    def test_search_endpoint(self, manager):
        try:
            from fastapi.testclient import TestClient
            from ctm_sak.viz.grafana.server import build_app
            client  = TestClient(build_app(manager))
            resp    = client.post("/search", json={})
            assert resp.status_code == 200
            targets = resp.json()
            assert isinstance(targets, list)
            assert any("test-ws" in t for t in targets)
        except ImportError:
            pytest.skip("fastapi not installed")

    def test_query_table_endpoint(self, manager):
        try:
            from fastapi.testclient import TestClient
            from ctm_sak.viz.grafana.server import build_app
            client = TestClient(build_app(manager))
            resp   = client.post("/query", json={
                "targets": [{"target": "test-ws/indicator", "type": "table"}],
                "range":   {"from": "now-30d", "to": "now"},
            })
            assert resp.status_code == 200
            results = resp.json()
            assert len(results) == 1
            assert "columns" in results[0]
            assert "rows" in results[0]
        except ImportError:
            pytest.skip("fastapi not installed")

    def test_query_summary_endpoint(self, manager):
        try:
            from fastapi.testclient import TestClient
            from ctm_sak.viz.grafana.server import build_app
            client = TestClient(build_app(manager))
            resp   = client.post("/query", json={
                "targets": [{"target": "test-ws/summary"}],
                "range":   {},
            })
            assert resp.status_code == 200
            results = resp.json()
            assert results[0]["type"] == "table"
            type_names = [row[0] for row in results[0]["rows"]]
            assert "indicator" in type_names
        except ImportError:
            pytest.skip("fastapi not installed")

    def test_query_timeseries_endpoint(self, manager):
        try:
            from fastapi.testclient import TestClient
            from ctm_sak.viz.grafana.server import build_app
            client = TestClient(build_app(manager))
            resp   = client.post("/query", json={
                "targets": [{"target": "test-ws/indicator/confidence"}],
                "range":   {},
            })
            assert resp.status_code == 200
            results = resp.json()
            assert "datapoints" in results[0]
        except ImportError:
            pytest.skip("fastapi not installed")

    def test_tag_keys_endpoint(self, manager):
        try:
            from fastapi.testclient import TestClient
            from ctm_sak.viz.grafana.server import build_app
            client = TestClient(build_app(manager))
            resp   = client.post("/tag-keys", json={})
            assert resp.status_code == 200
            keys   = [item["text"] for item in resp.json()]
            assert "stix_type" in keys
        except ImportError:
            pytest.skip("fastapi not installed")

    def test_tag_values_stix_type(self, manager):
        try:
            from fastapi.testclient import TestClient
            from ctm_sak.viz.grafana.server import build_app
            client = TestClient(build_app(manager))
            resp   = client.post("/tag-values", json={"key": "stix_type"})
            assert resp.status_code == 200
            values = [item["text"] for item in resp.json()]
            assert "indicator" in values
        except ImportError:
            pytest.skip("fastapi not installed")

    def test_annotations_endpoint(self, manager):
        try:
            from fastapi.testclient import TestClient
            from ctm_sak.viz.grafana.server import build_app
            client = TestClient(build_app(manager))
            resp   = client.post("/annotations", json={
                "annotation": {"query": "test-ws", "name": "Enrichment"}
            })
            assert resp.status_code == 200
            assert isinstance(resp.json(), list)
        except ImportError:
            pytest.skip("fastapi not installed")


# ===========================================================================
# GraphView — intent-driven rendering API
# ===========================================================================

class TestGraphViewIntents:
    """Tests for the five intent rendering methods."""

    def test_render_relationship_graph_has_edges(self, tmp_path):
        ws  = _populated_workspace(tmp_path)
        out = str(tmp_path / "rel.html")
        GraphView(ws).render_relationship_graph(path=out)
        content = Path(out).read_text()
        assert "GRAPH_DATA" in content
        # Extract edge count from embedded JSON
        import json as _json
        start = content.find("const GRAPH_DATA = ") + len("const GRAPH_DATA = ")
        end   = content.find(";\n\n    // ── Build", start)
        gd    = _json.loads(content[start:end])
        assert len(gd["edges"]) > 0, "relationship graph must have edges"

    def test_render_relationship_graph_filter(self, tmp_path):
        ws  = _populated_workspace(tmp_path)
        out = str(tmp_path / "rel_filter.html")
        # Filter to relationship type not present in test data → no edges
        GraphView(ws).render_relationship_graph(
            relationship_types=["attributed-to"], path=out
        )
        import json as _json
        content = Path(out).read_text()
        start = content.find("const GRAPH_DATA = ") + len("const GRAPH_DATA = ")
        end   = content.find(";\n\n    // ── Build", start)
        gd    = _json.loads(content[start:end])
        assert len(gd["edges"]) == 0

    def test_render_type_graph_creates_file(self, tmp_path):
        ws  = _populated_workspace(tmp_path)
        out = str(tmp_path / "type.html")
        GraphView(ws).render_type_graph(path=out)
        assert Path(out).exists()
        assert "GRAPH_DATA" in Path(out).read_text()

    def test_render_type_graph_no_edges(self, tmp_path):
        import json as _json
        ws  = _populated_workspace(tmp_path)
        out = str(tmp_path / "type_ne.html")
        GraphView(ws).render_type_graph(show_edges=False, path=out)
        content = Path(out).read_text()
        start = content.find("const GRAPH_DATA = ") + len("const GRAPH_DATA = ")
        end   = content.find(";\n\n    // ── Build", start)
        gd    = _json.loads(content[start:end])
        assert len(gd["edges"]) == 0

    def test_render_campaign_graph_is_subgraph(self, tmp_path):
        import json as _json
        ws  = _populated_workspace(tmp_path, n_indicators=10, n_malware=3,
                                    n_vulns=3, n_rels=8)
        gv  = GraphView(ws)
        all_nodes, _ = gv._extract_graph()
        out = str(tmp_path / "camp.html")
        gv.render_campaign_graph(depth=2, path=out)
        content = Path(out).read_text()
        start = content.find("const GRAPH_DATA = ") + len("const GRAPH_DATA = ")
        end   = content.find(";\n\n    // ── Build", start)
        gd    = _json.loads(content[start:end])
        assert 0 < len(gd["nodes"]) <= len(all_nodes)

    def test_render_campaign_graph_seed_present(self, tmp_path):
        ws   = _populated_workspace(tmp_path)
        ind  = _make_indicator("seed-target.com")
        ws.add(ind, mark_dirty=False)
        out  = str(tmp_path / "camp_seed.html")
        GraphView(ws).render_campaign_graph(seed_ids=[ind.id], depth=1, path=out)
        assert ind.id in Path(out).read_text()

    def test_render_campaign_graph_bad_seeds_fallback(self, tmp_path):
        ws  = _populated_workspace(tmp_path)
        out = str(tmp_path / "camp_bad.html")
        # Non-existent seed — should fall back to top-degree, not crash
        GraphView(ws).render_campaign_graph(
            seed_ids=["indicator--nonexistent-id"], depth=1, path=out
        )
        assert Path(out).exists()

    def test_render_timeline_graph_time_axis(self, tmp_path):
        import json as _json
        ws  = _populated_workspace(tmp_path, n_indicators=5, n_malware=2,
                                    n_vulns=2, n_rels=3)
        out = str(tmp_path / "time.html")
        GraphView(ws).render_timeline_graph(path=out)
        content = Path(out).read_text()
        start = content.find("const GRAPH_DATA = ") + len("const GRAPH_DATA = ")
        end   = content.find(";\n\n    // ── Build", start)
        gd    = _json.loads(content[start:end])
        xs = [n["x"] for n in gd["nodes"]]
        # Nodes with timestamps should span the 0-20 range
        assert max(xs) - min(xs) > 0

    def test_render_timeline_graph_custom_field(self, tmp_path):
        ws  = _populated_workspace(tmp_path)
        out = str(tmp_path / "time2.html")
        # Should not crash with a valid alternative time field
        GraphView(ws).render_timeline_graph(time_field="modified", path=out)
        assert Path(out).exists()

    def test_render_risk_heatmap_no_edges(self, tmp_path):
        import json as _json
        ws  = _populated_workspace(tmp_path)
        out = str(tmp_path / "risk.html")
        GraphView(ws).render_risk_heatmap(path=out)
        content = Path(out).read_text()
        start = content.find("const GRAPH_DATA = ") + len("const GRAPH_DATA = ")
        end   = content.find(";\n\n    // ── Build", start)
        gd    = _json.loads(content[start:end])
        assert len(gd["edges"]) == 0

    def test_render_risk_heatmap_value_positions(self, tmp_path):
        import json as _json
        ws  = _populated_workspace(tmp_path, n_indicators=5, n_malware=0,
                                    n_vulns=0, n_rels=0)
        out = str(tmp_path / "risk2.html")
        GraphView(ws).render_risk_heatmap(
            x_field="confidence", y_field="x_rf_risk_score", path=out
        )
        content = Path(out).read_text()
        start = content.find("const GRAPH_DATA = ") + len("const GRAPH_DATA = ")
        end   = content.find(";\n\n    // ── Build", start)
        gd    = _json.loads(content[start:end])
        # X values should vary since indicators have different confidence levels
        xs = [n["x"] for n in gd["nodes"]]
        assert max(xs) > min(xs)

    def test_render_risk_heatmap_custom_fields(self, tmp_path):
        ws  = _populated_workspace(tmp_path, n_indicators=3, n_malware=0,
                                    n_vulns=3, n_rels=0)
        out = str(tmp_path / "risk3.html")
        GraphView(ws).render_risk_heatmap(
            x_field="confidence", y_field="x_cvss_score",
            stix_types=["vulnerability"], path=out
        )
        assert Path(out).exists()

    def test_all_intents_produce_valid_html(self, tmp_path):
        """Smoke test: every intent method produces parseable HTML."""
        ws = _populated_workspace(tmp_path)
        gv = GraphView(ws)
        methods = [
            ("render_relationship_graph", {}),
            ("render_type_graph", {}),
            ("render_campaign_graph", {}),
            ("render_timeline_graph", {}),
            ("render_risk_heatmap", {}),
        ]
        for method_name, kwargs in methods:
            out = str(tmp_path / f"{method_name}.html")
            getattr(gv, method_name)(path=out, **kwargs)
            content = Path(out).read_text()
            assert "<!DOCTYPE html>" in content, f"{method_name}: not valid HTML"
            assert "GRAPH_DATA" in content, f"{method_name}: no GRAPH_DATA"

    def test_ego_subgraph_depth_1(self, tmp_path):
        ws = _populated_workspace(tmp_path)
        gv = GraphView(ws)
        all_nodes, all_edges = gv._extract_graph()
        seeds = list(gv._top_by_degree(all_nodes, all_edges, 1).keys())
        ego_n, ego_e = gv._ego_subgraph(all_nodes, all_edges, seeds, depth=1)
        # At depth 1, ego should contain seed + direct neighbours only
        assert seeds[0] in ego_n

    def test_timeline_layout_missing_timestamps(self, tmp_path):
        ws = _populated_workspace(tmp_path)
        gv = GraphView(ws)
        # Add an indicator with no created field
        from ctm_sak.orm.indicator import Indicator as _Ind
        bare = _Ind(name="bare.com", pattern="[domain-name:value = 'bare.com']",
                    pattern_type="stix")
        ws.add(bare, mark_dirty=False)
        all_nodes, _ = gv._extract_graph()
        pos = gv._timeline_layout(all_nodes, "created")
        # bare indicator should get x = -5
        assert pos[bare.id][0] == -5.0

    def test_risk_layout_missing_fields(self, tmp_path):
        ws = _populated_workspace(tmp_path)
        gv = GraphView(ws)
        # Add an indicator with no x_rf_risk_score
        from ctm_sak.orm.indicator import Indicator as _Ind
        bare = _Ind(name="bare2.com", pattern="[domain-name:value = 'bare2.com']",
                    pattern_type="stix", confidence=50)
        ws.add(bare, mark_dirty=False)
        all_nodes, _ = gv._extract_graph()
        pos = gv._risk_layout(all_nodes, "confidence", "x_rf_risk_score")
        # bare indicator has confidence but no x_rf_risk_score — y should be jittered
        assert bare.id in pos
